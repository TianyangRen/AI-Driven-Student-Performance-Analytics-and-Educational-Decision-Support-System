"""
文件解析层：把上传的文件字节 -> 规范化的行列表。

只负责「读成一行行 dict」，不做业务校验（那是 services 的事）。
按扩展名分流：.csv 用标准库，.xlsx 用 openpyxl。
"""
import csv
import io
from typing import Dict, List

from openpyxl import load_workbook


class ParseError(Exception):
    """解析层错误基类。"""


class UnsupportedFormat(ParseError):
    pass


class MissingColumns(ParseError):
    def __init__(self, missing: List[str]):
        self.missing = missing
        super().__init__(f"缺少必填列: {', '.join(missing)}")


class ParsedRow:
    """一行数据 + 其在原文件中的 1-based 行号（含表头，数据首行为 2）。"""

    def __init__(self, row_no: int, values: Dict[str, str]):
        self.row_no = row_no
        self.values = values


def _check_header(header: List[str], required: List[str]) -> None:
    present = {(h or "").strip() for h in header}
    missing = [c for c in required if c not in present]
    if missing:
        raise MissingColumns(missing)


def _is_blank(values: Dict[str, str]) -> bool:
    return all((v is None or str(v).strip() == "") for v in values.values())


def parse(file, filename: str, required_columns: List[str]) -> List[ParsedRow]:
    """
    Parameters
    ----------
    file : 类文件对象（request.FILES 的项）
    filename : 原文件名，用于判断扩展名
    required_columns : 必填列名，用于列头校验

    Returns
    -------
    List[ParsedRow] —— 跳过全空行；行号为原文件 1-based（数据首行=2）。

    Raises
    ------
    UnsupportedFormat / MissingColumns
    """
    lower = filename.lower()
    if lower.endswith(".csv"):
        return _parse_csv(file, required_columns)
    if lower.endswith(".xlsx"):
        return _parse_xlsx(file, required_columns)
    raise UnsupportedFormat(f"不支持的文件类型: {filename}（仅支持 .csv / .xlsx）")


def _parse_csv(file, required_columns: List[str]) -> List[ParsedRow]:
    raw = file.read()
    if isinstance(raw, bytes):
        # 兼容带 BOM 的 Excel 导出 CSV
        text = raw.decode("utf-8-sig")
    else:
        text = raw
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        raise MissingColumns(required_columns)
    header = [h.strip() for h in rows[0]]
    _check_header(header, required_columns)
    out: List[ParsedRow] = []
    for idx, raw_row in enumerate(rows[1:], start=2):
        values = {header[i]: (raw_row[i] if i < len(raw_row) else "") for i in range(len(header))}
        if _is_blank(values):
            continue
        out.append(ParsedRow(idx, values))
    return out


def _parse_xlsx(file, required_columns: List[str]) -> List[ParsedRow]:
    wb = load_workbook(io.BytesIO(file.read()), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_raw = next(rows_iter)
    except StopIteration:
        raise MissingColumns(required_columns)
    header = [(str(h).strip() if h is not None else "") for h in header_raw]
    _check_header(header, required_columns)
    out: List[ParsedRow] = []
    for idx, raw_row in enumerate(rows_iter, start=2):
        values = {
            header[i]: ("" if i >= len(raw_row) or raw_row[i] is None else str(raw_row[i]).strip())
            for i in range(len(header))
        }
        if _is_blank(values):
            continue
        out.append(ParsedRow(idx, values))
    wb.close()
    return out

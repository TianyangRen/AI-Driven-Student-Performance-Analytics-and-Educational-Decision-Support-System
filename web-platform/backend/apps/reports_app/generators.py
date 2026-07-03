"""报告生成 —— 用 openpyxl 生成真实 xlsx 落盘到 settings.REPORT_OUTPUT_DIR。

三种报告：
  CLASS_SUMMARY  班级总览：KPI + 学生明细（均分/出勤/风险）
  RISK_LIST      风险名单：按最近一次预测运行的挂科概率降序
  COMPARISON     对比分析：维度 × 考核类型的平均得分率矩阵

数据全部来自 apps.analytics.services（与页面接口同源，口径一致）。
"""
from __future__ import annotations

import os
from statistics import mean, median

from django.conf import settings
from openpyxl import Workbook

from apps.analytics import services
from apps.courses.models import Enrollment, Student


def generate(report) -> str:
    """按 report.report_type 生成文件，返回绝对路径。"""
    builders = {
        "CLASS_SUMMARY": _class_summary,
        "RISK_LIST": _risk_list,
        "COMPARISON": _comparison,
    }
    builder = builders.get(report.report_type)
    if builder is None:
        raise ValueError(f"不支持的报告类型: {report.report_type}")
    wb = builder(report)

    os.makedirs(settings.REPORT_OUTPUT_DIR, exist_ok=True)
    path = os.path.join(
        settings.REPORT_OUTPUT_DIR,
        f"report_{report.id}_{report.report_type.lower()}.xlsx",
    )
    wb.save(path)
    return path


def filename_for(report) -> str:
    return f"report_{report.id}_{report.report_type.lower()}.xlsx"


def _class_summary(report) -> Workbook:
    section = report.section
    ids = services.active_student_ids(section)
    scores = services.student_scores(section, ids)
    attendance = services.attendance_rates(section, ids)
    rmap, run = services.risk_by_student(section)
    vals = list(scores.values())

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Class Summary", f"{section.course.code}-{section.section_code}"])
    ws.append(["Course", section.course.name])
    ws.append(["Term", section.course.term])
    ws.append([])
    ws.append(["Students", len(ids)])
    ws.append(["Average score", round(mean(vals), 1) if vals else 0])
    ws.append(["Median score", round(median(vals), 1) if vals else 0])
    ws.append(["Pass rate (%)", round(sum(1 for v in vals if v >= services.PASS_MARK) / len(vals) * 100, 1) if vals else 0])
    rs = services.risk_summary(section, ids)
    ws.append(["Risk high / medium / low", f"{rs['high']} / {rs['medium']} / {rs['low']}"])

    ws2 = wb.create_sheet("Students")
    ws2.append(["Student", "Average", "Attendance (%)", "Risk level"])
    enrollments = (
        Enrollment.objects.filter(section=section, status="ACTIVE")
        .select_related("student")
    )
    for e in enrollments:
        ws2.append([
            e.student.anonymized_code,
            round(scores.get(e.student_id, 0.0), 1),
            round(attendance.get(e.student_id, 0.0) * 100, 0),
            (rmap.get(e.student_id) or {}).get("level", "LOW"),
        ])
    return wb


def _risk_list(report) -> Workbook:
    section = report.section
    rmap, run = services.risk_by_student(section)
    students = {s.id: s for s in Student.objects.filter(id__in=list(rmap.keys()))}

    wb = Workbook()
    ws = wb.active
    ws.title = "Risk List"
    ws.append(["Risk List", f"{section.course.code}-{section.section_code}"])
    ws.append(["Prediction run", run.created_at.isoformat() if run else "N/A"])
    ws.append([])
    ws.append(["Student", "Risk level", "Fail probability"])
    for sid, info in sorted(rmap.items(), key=lambda kv: -kv[1]["probability"]):
        student = students.get(sid)
        ws.append([
            student.anonymized_code if student else str(sid),
            info["level"],
            round(info["probability"], 4),
        ])
    if not rmap:
        ws.append(["(No successful prediction run yet — run predictions first.)"])
    return wb


def _comparison(report) -> Workbook:
    dimension = (report.parameters.get("dimension") or "SECTION").upper()
    data = services.comparison(report.requested_by, dimension)

    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"
    ws.append(["Comparison", dimension])
    ws.append([])
    ws.append([""] + list(data["labels"]))
    for s in data["series"]:
        ws.append([s["name"]] + list(s["data"]))
    return wb
